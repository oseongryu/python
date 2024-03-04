import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet('Sheet')

ws['A1'] = '참가번호'
ws['B1'] = '참가번호'

ws['A2'] = 1
ws['B2'] = '테스트'

wb.save('./test.xlsx')

wb = openpyxl.load_workbook('./test.xlsx')

ws = wb['Sheet']

ws['A3'] = 55
ws['B3'] = '테스트트'
wb.save('./test.xlsx')
